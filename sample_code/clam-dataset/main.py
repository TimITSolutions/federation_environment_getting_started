import os
import sys
import math
import torch
import random
import numpy as np
import pandas as pd
import openpyxl as op
import torch.nn as nn
from PIL import Image
from sklearn import metrics
import torch.optim as optim
from openpyxl import Workbook
import torch.nn.functional as F
from torchvision import transforms
from torch.utils.data import Dataset
from torch.utils.data import DataLoader
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from torchvision.models import resnet18, ResNet18_Weights

import mlflow

def set_random_seed(config_params):
    #random state initialization of the code - values - 8, 24, 30
    torch.manual_seed(config_params['randseedother']) 
    torch.cuda.manual_seed(config_params['randseedother'])
    torch.cuda.manual_seed_all(config_params['randseedother'])
    np.random.seed(config_params['randseeddata'])
    random.seed(config_params['randseeddata'])
    g = torch.Generator()
    g.manual_seed(config_params['randseedother'])
    torch.backends.cudnn.deterministic = True
    return g

def seed_worker(worker_id):
    worker_seed = torch.initial_seed() % 2**32
    np.random.seed(worker_seed)
    random.seed(worker_seed)

class BreastCancerDataset_generator(Dataset):
    def __init__(self, config_params, df, transform=None):
        """
        Args:
            csv_file (string): Path to the csv file with annotations.
            root_dir (string): Directory with all the images.
            transform (callable, optional): Optional transform to be applied
                on a sample.
        """
        self.df = df
        self.transform = transform
        self.config_params = config_params

    def __len__(self):
        return self.df.shape[0]
    
    def __getitem__(self, idx):
        data = self.df.iloc[idx]
        img = collect_images(self.config_params, data)
        img = self.transform(img)
        img = img.unsqueeze(0)
        return idx, img, torch.tensor(self.config_params['groundtruthdic'][data['Groundtruth']]), data['Views']

def collect_images(config_params, data):
    if config_params['bitdepth'] ==  8:
        img = collect_images_8bits(data)
    elif config_params['bitdepth'] == 12:
        img = collect_images_12bits(data)
    return img

def collect_images_8bits(data):
    img_path = str(data['FullPath'])
    img = Image.open(img_path)
    img= img.convert('RGB')
    transformTen = transforms.ToTensor()
    img = transformTen(img)
    return img

def collect_images_12bits(data):
    img_path = str(data['FullPath'])
    img = np.load(img_path).astype(np.float32)
    img = np.repeat(img[:, :, np.newaxis], 3, axis=2)
    img/=4095
    img = torch.from_numpy(img.transpose((2, 0, 1))).contiguous()
    return img

def save_model(model, optimizer, epoch, loss, path_to_model):
    state = {
        'epoch': epoch+1,
        'state_dict': model.state_dict(),
        'optim_dict': optimizer.state_dict(),
        'loss': loss
    }
    torch.save(state, path_to_model)

def load_model_for_testing(model, path):
    checkpoint = torch.load(path)
    model.load_state_dict(checkpoint['state_dict'])
    print("checkpoint epoch and loss:", checkpoint['epoch'], checkpoint['loss'])
    return model 

def results_store_excel(correct_train, total_images_train, train_loss, epoch, conf_mat_train, lr, path_to_results):
    lines = [epoch+1, lr]
    avg_train_loss=train_loss/total_images_train
    accuracy_train=correct_train / total_images_train
    speci_train=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
    recall_train=conf_mat_train[1,1]/sum(conf_mat_train[1,:])
    prec_train=conf_mat_train[1,1]/sum(conf_mat_train[:,1])
    f1_train=2*recall_train*prec_train/(recall_train+prec_train)
    prec_train_neg=conf_mat_train[0,0]/sum(conf_mat_train[:,0])
    recall_train_neg=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
    f1_train_neg=2*recall_train_neg*prec_train_neg/(recall_train_neg+prec_train_neg)
    f1macro_train=(f1_train+f1_train_neg)/2
    lines.extend([avg_train_loss, accuracy_train, f1macro_train, recall_train, speci_train])
    metrics_dic_train = {"loss_train": avg_train_loss, "accuracy_train": accuracy_train, "f1macro_train": f1macro_train}
    mlflow.log_metrics(metrics_dic_train, step=epoch)
    write_results_xlsx(path_to_results, 'train_val_results', lines)

def conf_mat_create(predicted, true, correct, conf_mat, classes):
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat_batch=confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    conf_mat=conf_mat+conf_mat_batch
    return correct, conf_mat, conf_mat_batch

def aggregate_performance_metrics(y_true, y_pred, y_prob): 
    prec_bin = metrics.precision_score(y_true, y_pred, pos_label = 1, average = 'binary')
    precmicro = metrics.precision_score(y_true, y_pred, average = 'micro')
    precmacro = metrics.precision_score(y_true, y_pred, average = 'macro')
    
    recall_bin = metrics.recall_score(y_true, y_pred, pos_label = 1, average = 'binary')
    recallmicro = metrics.recall_score(y_true, y_pred, average = 'micro')
    recallmacro = metrics.recall_score(y_true, y_pred, average = 'macro')
    
    f1_bin = metrics.f1_score(y_true, y_pred, pos_label = 1, average = 'binary')
    f1micro = metrics.f1_score(y_true, y_pred, average = 'micro')
    f1macro = metrics.f1_score(y_true, y_pred, average='macro')
    f1wtmacro = metrics.f1_score(y_true, y_pred, average='weighted')
    
    acc = metrics.accuracy_score(y_true, y_pred)
    cohen_kappa = metrics.cohen_kappa_score(y_true, y_pred)
    auc = metrics.roc_auc_score(y_true,y_prob)
    
    each_model_metrics = [prec_bin, precmicro, precmacro, recall_bin, recallmicro, recallmacro, f1_bin, f1micro, f1macro, f1wtmacro, acc, cohen_kappa, auc]
    return each_model_metrics

def classspecific_performance_metrics(config_params, y_true, y_pred):
    score_dict = classification_report(y_true, y_pred, labels=config_params['classes'], output_dict = True)
    #print(score_dict)
    results_all = []
    flag=0
    for key in score_dict.keys():
        if isinstance(score_dict[key], dict):
            if flag == 0:
                results_all.append(['class'] + list(score_dict[key].keys()))
                flag = 1
            results_all.append([key] + list(score_dict[key].values())) 
        else:
            results_all.append([key, score_dict[key]])
    
    print("class specific performance:", results_all)
    return results_all

def write_results_classspecific(path_to_results, sheetname, results_all):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    for result in results_all:
        sheet.append(result)
    wb.save(path_to_results)

def write_results_xlsx(path_to_results, sheetname, results):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    sheet.append(results)
    wb.save(path_to_results)  

def MyCollate(batch):
    i=0
    index=[]
    target=[]
    for item in batch:
        if i==0:
            data = batch[i][1]
            views_names = [item[3]]
        else:
            data=torch.cat((data,batch[i][1]), dim=0)
            views_names.append(item[3])
        index.append(item[0])
        target.append(item[2])
        i+=1
    index = torch.LongTensor(index)
    target = torch.LongTensor(target)
    return [index, data, target, views_names]

def train(dataloader_train, numbatches_train, model):
    for epoch in range(0, config_params['maxepochs']):
        model.train()
        loss_train=0.0
        batch_no=0
        correct_train = 0
        conf_mat_train=np.zeros((config_params['numclasses'],config_params['numclasses']))
        for train_idx, train_batch, train_labels, views_names in dataloader_train:
            train_batch = train_batch.to(config_params['device'])
            train_labels = train_labels.to(config_params['device'])
            train_labels = train_labels.view(-1)
            
            optimizer.zero_grad()

            output_batch = model(train_batch)
            loss = criterion(output_batch, train_labels)
            loss_batch=loss.item()
            loss.backward()
            optimizer.step()

            pred = output_batch.argmax(dim=1, keepdim=True)
            correct_train, conf_mat_train, _ = conf_mat_create(pred, train_labels, correct_train, conf_mat_train, config_params['classes'])
            loss_train+=(train_labels.size()[0]*loss.item())
            batch_no=batch_no+1
            print('Train: Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}'.format(epoch+1, config_params['maxepochs'], batch_no, numbatches_train, loss_batch), flush = True)

        #save model
        running_train_loss = loss_train/train_instances
        save_model(model, optimizer, epoch, running_train_loss, path_to_model)
        results_store_excel(correct_train, train_instances, loss_train, epoch, conf_mat_train, config_params['lr'], path_to_results)

def test(path_to_model, dataloader_test, numbatches_test, model):
    model = load_model_for_testing(model, path_to_model)
    model.eval()
    batch_test_no=0
    test_loss = 0
    for test_idx, test_batch, test_labels, views_names in dataloader_test:
        test_batch, test_labels = test_batch.to(config_params['device']), test_labels.to(config_params['device'])
        test_labels = test_labels.view(-1)
        output_test = model(test_batch)
        test_pred = output_test.argmax(dim=1, keepdim=True)
        loss = criterion(output_test, test_labels).item()
        
        test_loss += test_labels.size()[0]*loss
        
        if batch_test_no==0:
            test_pred_all = test_pred
            test_labels_all = test_labels
            output_all_ten = F.softmax(output_test.data,dim=1)
            output_all_ten = output_all_ten[:,1]
        else:
            test_pred_all = torch.cat((test_pred_all,test_pred),dim=0)
            test_labels_all = torch.cat((test_labels_all,test_labels),dim=0)
            output_all_ten=torch.cat((output_all_ten,F.softmax(output_test.data,dim=1)[:,1]),dim=0)
    
        batch_test_no+=1
        print('Test: Step [{}/{}], Loss: {:.4f}'.format(batch_test_no, numbatches_test, loss), flush=True)
    
    running_loss_test = test_loss/test_instances

    # save metrics on test set
    per_model_metrics = aggregate_performance_metrics(test_labels_all.cpu().numpy(), test_pred_all.cpu().numpy(), output_all_ten.cpu().numpy())
    per_model_metrics = [running_loss_test] + per_model_metrics
    print('Loss','PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC')
    print("scores:", per_model_metrics)
    write_results_xlsx(path_to_results, 'test_results', per_model_metrics)
    results_classspecific = classspecific_performance_metrics(config_params, test_labels_all.cpu().numpy(), test_pred_all.cpu().numpy())
    write_results_classspecific(path_to_results, 'test_results', results_classspecific)

    # mlflow metrics
    metrics_dic = {"Test loss": per_model_metrics[0], "Test precision positive class": per_model_metrics[1], "Test precision micro": per_model_metrics[2], "Test precision macro": per_model_metrics[3], "Test recall positive class": per_model_metrics[4], "Test recall micro": per_model_metrics[5], "Test recall macro": per_model_metrics[6], "Test f1 positive class": per_model_metrics[7], "Test f1 micro": per_model_metrics[8], "Test f1 macro": per_model_metrics[9], "Test f1 wt macro": per_model_metrics[10], "Test accuracy": per_model_metrics[11], "Test cohen kappa": per_model_metrics[12], "Test AUC": per_model_metrics[13]}
    mlflow.log_metrics(metrics_dic)

if __name__=='__main__':
    # input parameters for the code
    config_params = {'randseeddata': 8,
                    'randseedother': 8, 
                    'csvfilepath_image': '/mnt/dataset/clam-details-image.csv',
                    'preprocessed_imagepath': '/mnt/dataset',
                    'device': 'cuda:0', # cuda:0, cpu
                    'maxepochs': 2,
                    'path_to_output': '/mnt/export',
                    'batchsize': 2,
                    'resize':[224, 224],
                    'numclasses': 2,
                    'classes': [0, 1],
                    'numworkers': 0,
                    'bitdepth': 8, # change to 12 for using 12 bit .npy files.
                    'lr': 0.00001,
                    'wtdecay':0.0001,
                    'groundtruthdic': {'benign': 0, 'malignant': 1}}

    # Create a new MLflow Experiment
    # set username and password through environment variables. 
    # This is needed for accessing the mlflow client when submitting your code to fe.zgt.nl.
    username = 'your-username' #'your-username' #can be found when logging to fe.zgt.nl
    password = 'your-password' #'your-password' #can be found when logging to fe.zgt.nl
    
    os.environ["MLFLOW_TRACKING_USERNAME"] = username
    os.environ["MLFLOW_TRACKING_PASSWORD"] = password
    
    # Set our tracking server uri for logging
    mlflow.set_tracking_uri(uri="http://localhost:3001")
    
    mlflow.set_experiment(username)

    mlflow.start_run()

    #with mlflow.start_run():
    mlflow.log_params(config_params)


    # set random seed
    g = set_random_seed(config_params)

    print_dir = os.path.join(config_params['path_to_output'],'out.txt')
    tqdm_dir = os.path.join(config_params['path_to_output'],'error.txt')
    sys.stdout.close()
    sys.stderr.close()
    sys.stdout = open(print_dir, 'w')
    sys.stderr = open(tqdm_dir, 'w')
    

    # read input csv files
    csv_file_path = config_params['csvfilepath_image']
    df_modality = pd.read_csv(csv_file_path, sep=';')
    print("df modality shape:", df_modality.shape)
    df_modality['FullPath'] = config_params['preprocessed_imagepath'] + '/' + df_modality['ImagePath']
    df_modality['Groundtruth'] = df_modality['CaseLabel']

    # merging of image csv file with case csv file to get the BIRADS score, breast density and age information for each image. 
    # This is needed if you want to test on certain subgroups of the data, e.g. breast density A, B, C, D or
    # birads 0,1,2,3,4,4a,4b,4c,5,6
    df_case = pd.read_csv("/mnt/dataset/clam-details-case-extrainfo.csv", sep=';')
    print("dataframe shape:", df_modality.shape)
    df_modality = df_modality.merge(df_case, on='CaseName', how='inner')
    df_modality = df_modality.drop(['Patient_Id_y', 'CasePath_y', 'Study_Description_y', 'Views_y', 'Groundtruth_y', 'Split_y'], axis=1)
    df_modality = df_modality.rename(columns={'Patient_Id_x': 'Patient_Id', 'CasePath_x': 'CasePath', 'Study_Description_x': 'Study_Description', 'Views_x': 'Views', 'Split_x':'Split', 'Groundtruth_x': 'Groundtruth'})
    print("test set shape after merging with clam-details-case-extrainfo:", df_modality.shape)
    print("dataframe columns:", df_modality.columns)

    # result output
    path_to_results = config_params['path_to_output'] + '/' + 'results.xlsx'
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "train_val_results"
    header = ['Epoch','lr','Avg Loss Train','Accuracy Train','F1macro Train','Recall Train','Speci Train']
    sheet1.append(header)
    sheet2 = wb.create_sheet('test_results') 
    sheet2.append(['Loss','PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC'])
    wb.save(path_to_results)

    # saved model path
    path_to_model = config_params['path_to_output'] + '/' + 'model.tar'

    # data augmentation
    mean = [0.485, 0.456, 0.406]
    std_dev = [0.229, 0.224, 0.225]
    preprocess_train = transforms.Compose([transforms.Resize((config_params['resize'][0], config_params['resize'][1])),
                                        transforms.RandomHorizontalFlip(p=0.5),
                                        transforms.RandomAffine(degrees=(15),translate=(0.1,0.1),scale=(0.8,1.6),shear=(25)),
                                        transforms.Normalize(mean=mean, std=std_dev)])

    preprocess_test = transforms.Compose([transforms.Resize((config_params['resize'][0], config_params['resize'][1])),
                                                            transforms.Normalize(mean=mean, std=std_dev)])

    # dataloader for train and test set
    df_train = df_modality[df_modality['Split'] == 'train']
    df_test = df_modality[df_modality['Split'] == 'test']

    if df_train.shape[0]>100:
        df_train = df_train[100:150]
        df_test = df_test[0:70]

    df_train = df_train.reset_index()
    df_test = df_test.reset_index()
    train_instances = df_train.shape[0]
    test_instances = df_test.shape[0]
    print("train instances:", train_instances)
    print("test instances:", test_instances)

    dataset_gen_train = BreastCancerDataset_generator(config_params, df_train, preprocess_train)
    dataloader_train = DataLoader(dataset_gen_train, batch_size=config_params['batchsize'], shuffle=True, num_workers=config_params['numworkers'], collate_fn=MyCollate, worker_init_fn=seed_worker, generator=g)
    numbatches_train = int(math.ceil(train_instances/config_params['batchsize']))
                
    dataset_gen_test = BreastCancerDataset_generator(config_params, df_test, preprocess_test)
    dataloader_test = DataLoader(dataset_gen_test, batch_size=config_params['batchsize'], shuffle=False, num_workers=config_params['numworkers'], collate_fn=MyCollate, worker_init_fn=seed_worker, generator=g)
    numbatches_test = int(math.ceil(test_instances/config_params['batchsize']))

    # model initialization
    model_init = resnet18(weights=ResNet18_Weights.DEFAULT, progress=False)
    num_ftrs = model_init.fc.in_features
    model_init.fc = nn.Linear(num_ftrs, 2)
    model_init.to(config_params['device'])

    # loss function and optimizer
    criterion = nn.CrossEntropyLoss()
    optimizer = optim.Adam(filter(lambda p: p.requires_grad, model_init.parameters()), lr=config_params['lr'], weight_decay=config_params['wtdecay'])

    # model training
    train(dataloader_train, numbatches_train, model_init)

    # test model
    test(path_to_model, dataloader_test, numbatches_test, model_init)

    sys.stdout.close()
    sys.stderr.close()